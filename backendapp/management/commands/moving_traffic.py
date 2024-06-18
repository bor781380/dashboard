import decimal

import openpyxl
from django.core.management.base import BaseCommand
from backendapp.models import Traffic, Shops

class Command(BaseCommand):
    help = "Перенос данных из Excel в базу данных"

    def handle(self, *args, **options):
        # Очищаем таблицу
        Traffic.objects.all().delete()
        # Открываем Excel-файл
        workbook = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\traffic.xlsx')
        sheet = workbook.active

        # Перенос данных из Excel в базу данных
        for row in range(2, sheet.max_row + 1):
            # Читаем данные из первого столбца Excel
            shop_name = sheet.cell(row=row, column=1).value
            APPG_value = sheet.cell(row=row, column=2).value
            count_value= sheet.cell(row=row, column=3).value
            planFact_value = sheet.cell(row=row, column=4).value

            # Конвертируем значение пустых из строки в число, если оно не пустое
            if APPG_value:
                APPG = decimal.Decimal(APPG_value)
                count = int(count_value)
                planFact = decimal.Decimal(planFact_value)


            else:
                APPG = 0.0
                count = 0.0
                planFact = 0.0
            # Создаем или получаем экземпляр модели Shops
            shop, _ = Shops.objects.get_or_create(name=shop_name)

            # Создаем новый объект модели и сохраняем его
            traffic = Traffic(shops=shop, APPG=APPG, planFact=planFact, count=count)

            traffic.save()

        self.stdout.write(self.style.SUCCESS("Данные успешно перенесены в базу данных!"))