import decimal
import openpyxl
from datetime import date
from django.core.management.base import BaseCommand
from backendapp.models import ChartSpecialTask, Shops

class Command(BaseCommand):
    help = "Перенос данных из Excel в базу данных"

    def handle(self, *args, **options):
        # Очищаем таблицу
        ChartSpecialTask.objects.all().delete()
        # Открываем Excel-файл
        workbook = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\ChartSpecialTask.xlsx')
        sheet = workbook.active

        # Получаем общее количество записей в Excel
        total_rows = sheet.max_row - 1  # Вычитаем 1, чтобы исключить заголовок

        # Разбиваем данные на порции по 500 записей
        chunk_size = 500
        num_chunks = total_rows // chunk_size
        last_chunk_size = total_rows % chunk_size

        for chunk_index in range(num_chunks):
            start_row = chunk_index * chunk_size + 2  # Начинаем с 2-й строки (пропускаем заголовок)
            end_row = (chunk_index + 1) * chunk_size + 1

            data = []
            for row in range(start_row, end_row):
                shop_name = sheet.cell(row=row, column=1).value
                specialTask = sheet.cell(row=row, column=5).value
                date = sheet.cell(row=row, column=4).value
                sales_value = sheet.cell(row=row, column=2).value
                planFact_value = sheet.cell(row=row, column=9).value
                planDay = sheet.cell(row=row, column=7).value

                if sales_value:
                    sales = decimal.Decimal(sales_value)
                else:
                    sales = decimal.Decimal(0.00)

                if planFact_value:
                    planFact = decimal.Decimal(planFact_value)
                else:
                    planFact = decimal.Decimal(0.00)

                shop, _ = Shops.objects.get_or_create(name=shop_name)
                data.append(ChartSpecialTask(shops=shop,
                                             specialTask=specialTask,
                                             date=date,
                                             sales=sales,
                                             planFact=planFact,
                                             planDay=planDay))

            ChartSpecialTask.objects.bulk_create(data)
            self.stdout.write(self.style.SUCCESS(f"Записи с {start_row} по {end_row-1} успешно перенесены в базу данных."))

        # Обрабатываем последний блок данных, если он меньше 500 записей
        if last_chunk_size > 0:
            start_row = (num_chunks * chunk_size) + 2
            end_row = total_rows + 2

            data = []
            for row in range(start_row, end_row):
                shop_name = sheet.cell(row=row, column=1).value
                specialTask = sheet.cell(row=row, column=5).value
                date = sheet.cell(row=row, column=4).value
                sales_value = sheet.cell(row=row, column=2).value
                planFact_value = sheet.cell(row=row, column=9).value
                planDay = sheet.cell(row=row, column=7).value

                if sales_value:
                    sales = decimal.Decimal(sales_value)
                else:
                    sales = decimal.Decimal(0.00)

                if planFact_value:
                    planFact = decimal.Decimal(planFact_value)
                else:
                    planFact = decimal.Decimal(0.00)

                shop, _ = Shops.objects.get_or_create(name=shop_name)
                data.append(ChartSpecialTask(shops=shop,
                                             specialTask=specialTask,
                                             date=date,
                                             sales=sales,
                                             planFact=planFact,
                                             planDay=planDay))

            if data:
                ChartSpecialTask.objects.bulk_create(data)
                self.stdout.write(
                    self.style.SUCCESS(f"Записи с {start_row} по {end_row - 1} успешно перенесены в базу данных."))

            self.stdout.write(self.style.SUCCESS(f"Все {total_rows} записей успешно перенесены в базу данных!"))

