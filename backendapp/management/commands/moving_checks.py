import decimal

import openpyxl
from django.core.management.base import BaseCommand
from backendapp.models import Checks, Shops

class Command(BaseCommand):
    help = "Перенос данных из Excel в базу данных"

    def handle(self, *args, **options):
        # Очищаем таблицу
        Checks.objects.all().delete()
        # Открываем Excel-файл
        workbook = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\checks.xlsx')
        sheet = workbook.active

        # Перенос данных из Excel в базу данных
        for row in range(2, sheet.max_row + 1):
            # Читаем данные из первого столбца Excel
            shop_name = sheet.cell(row=row, column=1).value
            APPG_value = sheet.cell(row=row, column=2).value
            checksRoz_value = sheet.cell(row=row, column=3).value
            checksEcom_value = sheet.cell(row=row, column=4).value
            partCheckRoz_value = sheet.cell(row=row, column=5).value

            # Конвертируем значение пустых из строки в число, если оно не пустое
            if APPG_value:
                APPG = decimal.Decimal(APPG_value)
            else:
                APPG = decimal.Decimal(0.0)

            if checksRoz_value:
                checksRoz = int(checksRoz_value)
            else:
                checksRoz = 0

            if checksEcom_value:
                checksEcom = int(checksEcom_value)
            else:
                checksEcom = 0

            if partCheckRoz_value:
                partCheckRoz = decimal.Decimal(partCheckRoz_value)
            else:
                partCheckRoz = decimal.Decimal(0.0)
            # Создаем или получаем экземпляр модели Shops
            shop, _ = Shops.objects.get_or_create(name=shop_name)

            # Создаем новый объект модели и сохраняем его
            check = Checks(shops=shop, APPG=APPG, checksRoz=checksRoz, checksEcom=checksEcom, partCheckRoz=partCheckRoz)

            check.save()

        self.stdout.write(self.style.SUCCESS("Данные успешно перенесены в базу данных!"))