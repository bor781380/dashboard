import decimal
import openpyxl
from django.core.management.base import BaseCommand
from backendapp.models import SpecialTask, Shops

class Command(BaseCommand):
    help = "Перенос данных из Excel в базу данных"

    def handle(self, *args, **options):
        # Очищаем таблицу
        SpecialTask.objects.all().delete()
        # Открываем Excel-файл
        workbook = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\SpecialTask.xlsx')
        sheet = workbook.active

        COUNTTASK = 3

        for i in range(1, COUNTTASK + 1):
            # Перенос данных из Excel в базу данных
            for row in range(1, sheet.max_row):
                # Читаем данные из первого столбца Excel
                shop_name = sheet.cell(row=row+1, column=1).value
                name = sheet.cell(row=1, column= i + 1).value
                planFact_value = sheet.cell(row=row+1, column=i + 1).value

                # Конвертируем значение пустых из строки в число, если оно не пустое
                if planFact_value:
                    planFact = decimal.Decimal(planFact_value)
                else:
                    planFact = decimal.Decimal(0.0)

                # Создаем или получаем экземпляр модели Shops
                shop, _ = Shops.objects.get_or_create(name=shop_name)

                # Создаем новый объект модели и сохраняем его
                task = SpecialTask(shops=shop, name=name, planFact=planFact)

                task.save()

        self.stdout.write(self.style.SUCCESS("Данные успешно перенесены в базу данных!"))