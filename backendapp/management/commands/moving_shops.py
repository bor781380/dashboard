import openpyxl
from django.core.management.base import BaseCommand
from backendapp.models import Shops

class Command(BaseCommand):
    help = "Перенос данных из Excel в базу данных"

    def handle(self, *args, **options):
        # Очищаем таблицу
        Shops.objects.all().delete()
        # Открываем Excel-файл
        workbook = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\shops.xlsx')
        sheet = workbook.active

        # Перенос данных из Excel в базу данных
        for row in range(2, sheet.max_row + 1):
            # Читаем данные из первого столбца Excel
            column1 = sheet.cell(row=row, column=1).value

            # Создаем новый объект модели и сохраняем его
            shop = Shops(name=column1)
            shop.save()

        self.stdout.write(self.style.SUCCESS("Данные успешно перенесены в базу данных!"))