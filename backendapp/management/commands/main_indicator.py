import openpyxl

from django.core.management.base import BaseCommand
from backendapp.models import MainIndicator


def color_(value):
    col = 0
    if value < 95:
        col = 3
    elif value >= 100:
        col = 1
    else:
        col = 2
    return col

class Command(BaseCommand):
    help = "Формирование базы итогового показателя"

    def handle(self, *args, **options):
        # Очищаем таблицу
        MainIndicator.objects.all().delete()
        # Открываем Excel-файл
        workbookPROD = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\ProductLine.xlsx')
        sheetPROD = workbookPROD.active

        workbookSPEC = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\SpecialTask.xlsx')
        sheetSPEC = workbookSPEC.active

        workbookTRAF = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\traffic.xlsx')
        sheetTRAF = workbookTRAF.active

        workbookCHECK = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\checks.xlsx')
        sheetCHECK = workbookCHECK.active

        workbookCONV = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\Convers.xlsx')
        sheetCONV = workbookCONV.active

        data = []
        color_value = sheetPROD.cell(row=16, column=67).value
        planFactMI = sheetPROD.cell(row=16, column=68).value
        colorPROD = color_(planFactMI)
        marginMI = sheetPROD.cell(row=16, column=69).value
        colorNAC = color_(marginMI)
        balancesMI = sheetPROD.cell(row=16, column=70).value
        colorOST = color_(balancesMI)
        turnoverMI = sheetPROD.cell(row=16, column=71).value
        colorOBOR = color_(turnoverMI)
        colorSPEC = color_((sheetSPEC.cell(row=15, column=2).value +
                           sheetSPEC.cell(row=15, column=3).value +
                           sheetSPEC.cell(row=15, column=4).value)/3)
        colorTRAFF = color_(sheetTRAF.cell(row=15, column=2).value)
        colorCHECK = color_(sheetCHECK.cell(row=15, column=2).value)
        colorCONF = color_(sheetCONV.cell(row=15, column=2).value)

        data.append(MainIndicator(color_value=color_value,
                                  colorPROD=colorPROD,
                                  planFactMI=planFactMI,
                                  colorNAC=colorNAC,
                                  marginMI=marginMI,
                                  colorOST=colorOST,
                                  balancesMI=balancesMI,
                                  colorOBOR=colorOBOR,
                                  turnoverMI=turnoverMI,
                                  colorSPEC=colorSPEC,
                                  colorTRAFF=colorTRAFF,
                                  colorCHECK=colorCHECK,
                                  colorCONF=colorCONF))


        MainIndicator.objects.bulk_create(data)
        self.stdout.write(self.style.SUCCESS(f"Записи успешно перенесены в базу данных."))



