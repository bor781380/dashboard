import decimal
import openpyxl
from django.core.management.base import BaseCommand
from backendapp.models import ProductLine, Shops

class Command(BaseCommand):
    help = "Перенос данных из Excel в базу данных"

    def handle(self, *args, **options):
        # Очищаем таблицу
        ProductLine.objects.all().delete()
        # Открываем Excel-файл
        workbook = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\ProductLine.xlsx')
        sheet = workbook.active

        PRODUCTlINE_STEP = 5
        PRODUCTlINE = 14

        for i in range(1, PRODUCTlINE_STEP * PRODUCTlINE - 1, PRODUCTlINE_STEP):
        #for i in range(1, 2):
            # Перенос данных из Excel в базу данных
            for row in range(1, sheet.max_row-1):
                # Читаем данные из первого столбца Excel
                shop_name = sheet.cell(row=row+2, column=1).value
                name = sheet.cell(row=1, column= 1+i).value
                colorName = sheet.cell(row=2, column= 1+i).value
                color = sheet.cell(row=row+2, column=1+i).value
                planFactName = sheet.cell(row=2, column= 2+i).value
                planFact_value = sheet.cell(row=row+2, column=2+i).value
                marginName = sheet.cell(row=2, column= 3+i).value
                margin_value = sheet.cell(row=row+2, column=3+i).value
                balancesName = sheet.cell(row=2, column= 4+i).value
                balances_value = sheet.cell(row=row+2, column=4+i).value
                turnoverName = sheet.cell(row=2, column= 5+i).value
                turnover_value = sheet.cell(row=row+2, column=5+i).value


                # Конвертируем значение пустых из строки в число, если оно не пустое

                if planFact_value:
                    planFact = decimal.Decimal(planFact_value)
                else:
                    planFact = decimal.Decimal(0.0)

                if margin_value:
                    margin = decimal.Decimal(margin_value)
                else:
                    margin = decimal.Decimal(0.0)

                if balances_value:
                    balances = decimal.Decimal(balances_value)
                else:
                    balances = decimal.Decimal(0.0)

                if turnover_value:
                    turnover = decimal.Decimal(turnover_value)
                else:
                    turnover = decimal.Decimal(0.0)

                # Создаем или получаем экземпляр модели Shops
                shop, _ = Shops.objects.get_or_create(name=shop_name)

                # Создаем новый объект модели и сохраняем его
                product = ProductLine(shops=shop,
                                      name=name,
                                      colorName=colorName,
                                      color=color,
                                      planFactName=planFactName,
                                      planFact=planFact,
                                      marginName=marginName,
                                      margin=margin,
                                      balancesName=balancesName,
                                      balances=balances,
                                      turnoverName=turnoverName,
                                      turnover=turnover)

                product.save()

        self.stdout.write(self.style.SUCCESS("Данные успешно перенесены в базу данных!"))