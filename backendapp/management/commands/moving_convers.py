import decimal

import openpyxl
from django.core.management.base import BaseCommand
from backendapp.models import Conversion, Shops

class Command(BaseCommand):
    help = "Перенос данных из Excel в базу данных"

    def handle(self, *args, **options):
        # Очищаем таблицу
        Conversion.objects.all().delete()
        # Открываем Excel-файл
        workbook = openpyxl.load_workbook('C:\\Users\\Andrey\\ProjectWEB\\Convers.xlsx')
        sheet = workbook.active

        # Перенос данных из Excel в базу данных
        for row in range(2, sheet.max_row + 1):
            # Читаем данные из первого столбца Excel
            shop_name = sheet.cell(row=row, column=1).value
            planFact_value = sheet.cell(row=row, column=2).value
            conversionFact_value = sheet.cell(row=row, column=3).value
            APPG_value = sheet.cell(row=row, column=4).value

            # Конвертируем значение пустых из строки в число, если оно не пустое
            if APPG_value:
                APPG = decimal.Decimal(APPG_value)
            else:
                APPG = decimal.Decimal(0.0)

            if conversionFact_value:
                conversionFact = decimal.Decimal(conversionFact_value)
            else:
                conversionFact = decimal.Decimal(0.0)


            if planFact_value:
                planFact = decimal.Decimal(planFact_value)
            else:
                planFact = decimal.Decimal(0.0)
            # Создаем или получаем экземпляр модели Shops
            shop, _ = Shops.objects.get_or_create(name=shop_name)

            # Создаем новый объект модели и сохраняем его
            convers = Conversion(shops=shop, planFact=planFact, conversionFact=conversionFact, APPG=APPG)

            convers.save()

        self.stdout.write(self.style.SUCCESS("Данные успешно перенесены в базу данных!"))